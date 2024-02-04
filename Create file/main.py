import openpyxl
from openpyxl.utils.cell import column_index_from_string
import re

''' Creating an excel file based on another excel file '''
wb = openpyxl.load_workbook("IPA 248 Admitad MTD 15th November'23 (1).xlsx")
sheet = wb['DATA']
wb.create_sheet(title='filtered rows')
filtered_rows = wb['filtered rows']
correct_value = re.compile(r'(\d{13})')
for row in range(2, sheet.max_row):
    if sheet.cell(row=row, column=column_index_from_string('S')).value:
        for groups in correct_value.findall(sheet.cell(row=row, column=column_index_from_string('S')).value):
            filtered_rows['S' + str(row)] = groups
            filtered_rows['O' + str(row)] = sheet.cell(row=row, column=column_index_from_string('O')).value
            filtered_rows['N' + str(row)] = sheet.cell(row=row, column=column_index_from_string('N')).value
            filtered_rows['B' + str(row)] = sheet.cell(row=row, column=column_index_from_string('B')).value
            filtered_rows['A' + str(row)] = sheet.cell(row=row, column=column_index_from_string('A')).value
            filtered_rows['E' + str(row)] = sheet.cell(row=row, column=column_index_from_string('E')).value
            filtered_rows['C' + str(row)] = sheet.cell(row=row, column=column_index_from_string('C')).value
            filtered_rows['Y' + str(row)] = sheet.cell(row=row, column=column_index_from_string('Y')).value
    else:
        pass  # пометить строку с пустым значением 'S28'
sheet = wb['filtered rows']
empty_rows = [i for i in range(2, sheet.max_row)
              if sheet.cell(row=i, column=column_index_from_string('S')).value is None]
for i in reversed(empty_rows):
    sheet.delete_rows(i)
result = openpyxl.load_workbook('postbacks-import-template (42).xlsx')
table = result['Example CSV']
for row in range(2, sheet.max_row):
    value_A = sheet.cell(row=row, column=column_index_from_string('A')).value
    value_S = sheet.cell(row=row, column=column_index_from_string('S')).value
    if sheet.cell(row=row, column=column_index_from_string('O')).value == 'Declined':
        table['C'+str(row)] = value_S
        table['G'+str(row)] = 'Declined'
        value_N = sheet.cell(row=row, column=column_index_from_string('N')).value
        table['H'+str(row)] = f'The customer with application number {value_A} was rejected due to {value_N}.'
    elif sheet.cell(row=row, column=column_index_from_string('O')).value == 'Approved':
        table['C' + str(row)] = value_S
        table['G' + str(row)] = 'Approved'
        value_E = sheet.cell(row=row, column=column_index_from_string('E')).value
        table['H' + str(row)] = f'The customer with application number {value_A} has got {value_E}.'
    elif sheet.cell(row=row, column=column_index_from_string('O')).value == 'IPA':
        if sheet.cell(row=row, column=column_index_from_string('Y')).value is None:
            if sheet.cell(row=row, column=column_index_from_string('C')).value == 'STPK':
                table['C' + str(row)] = value_S
                table['G' + str(row)] ='Pending'
                table['H' + str(row)] = f'The client with application number {value_A} has gotten initial approval ' \
                                        f'from the bank. The client has to complete KYC verification by using this ' \
                                        f'link: http://www.axisbank.com/vkyc. Ask the client to complete the KYC.'
            elif sheet.cell(row=row, column=column_index_from_string('C')).value == 'STPI':
                table['C' + str(row)] = value_S
                table['G' + str(row)] = 'Pending'
                table['H' + str(row)] = f'The client with application number {value_A} has gotten initial approval ' \
                                        f'from the bank. The client has to provide income proof. The bank will ' \
                                        f'contact the client'
            elif sheet.cell(row=row, column=column_index_from_string('C')).value == 'STPT':
                table['C' + str(row)] = value_S
                table['G' + str(row)] = 'Pending'
                table['H' + str(row)] = f'The client with application number {value_A} has gotten initial approval' \
                                        f' from the bank. The client has to provide income proof and KYC verification' \
                                        f' by using this link: http://www.axisbank.com/vkyc.'
            else:
                table['C' + str(row)] = value_S
                table['G' + str(row)] = 'Pending'
                table['H' + str(row)] = f'The client with application number {value_A} has gotten initial approval' \
                                        f' from the bank. The bank will contact the client.'
        elif sheet.cell(row=row, column=column_index_from_string('Y')).value == 'DROPOFF':
            table['C' + str(row)] = value_S
            table['G' + str(row)] = 'Pending'
            table['H' + str(row)] = f'The client with application number {value_A} has dropped the video KYC.' \
                                    f' Ask the client to complete it by using this link http://www.axisbank.com/vkyc.'
        else:
            table['C' + str(row)] = value_S
            table['G' + str(row)] = 'Pending'
            table['H' + str(row)] = f'The client with application number {value_A} has finished the video KYC.' \
                                    f' The client has to wait the final decision of the bank.'
    elif sheet.cell(row=row, column=column_index_from_string('O')).value == 'RCU' or 'U/W' or 'UW Completed':
        table['C' + str(row)] = value_S
        table['G' + str(row)] = 'Pending'
        table['H' + str(row)] = f'The client’s application is {value_A}. The Risk Team of the Bank is checking ' \
                                f'the profile, so your client has to wait.'
    elif sheet.cell(row=row, column=column_index_from_string('O')).value in \
            ('Audit', 'Rework', 'FI', 'Hunter', 'Multi-account case'):
        table['C' + str(row)] = value_S
        table['G' + str(row)] = 'Pending'
        table['H' + str(row)] = f'The client’s application is {value_A}. The bank is checking the profile,' \
                                f' so your client has to wait.'


result.save('test3.xlsx')



# if __name__ == '__main__':
