import openpyxl
import re

"""Phone number formats: 
phone_v1: 'x xxx xxx xxxx' or 'x-xxx-xxx-xx-xx'
phone_v2: ' xx-xx-xx '
phone_v3: ' xxx-xxx '
phone_v4: ' 071 735 93 96 '
phone_v5: ' 8(4152) 503-038 '
phone_v6: ' 0552 47 66 60 '
phone_v7: ' \d{12} '
phone_v8: '+7(965) 0-302-304 '
"""
phone_v1 = re.compile(r'''(
    (\d)
    (\s|-|\.)?           # delimiter
    (\d{3}|\(\d{3}\))  # region code
    (\s|-|\.)?           # delimiter
    (\d{3}) 
    (\s|-|\.)?           # delimiter   
    (\d{2})
    (\s|-|\.)?           # delimiter 
    (\d{2})
    )''', re.VERBOSE)
phone_v2 = re.compile(r'''(
    \D
    \s
    (\d{2})
    (\s|-)           # delimiter 
    (\d{2})
    (\s|-)           # delimiter 
    (\d{2})
    (\s|,)    
    )''', re.VERBOSE)
phone_v3 = re.compile(r'''(
    ([^)])
    \s
    (\d{3})
    (-)           # delimiter
    (\d{3})
    (\s|,)    
    )''', re.VERBOSE)
phone_v4 = re.compile(r'''(
    \D
    \s
    (\d{3})
    (\s|-|\.)           # delimiter 
    (\d{3})
    (\s|-|\.)           # delimiter 
    (\d{2})
    (\s|-|\.)           # delimiter
    (\d{2})
    (\s|,) 
    )''', re.VERBOSE)
phone_v5 = re.compile(r'''(
    \s
    (\d)
    (\s|-)?
    (\d{4}|\(\d{4}\))
    (\s|-)
    (\d{3})
    (\s|-)
    (\d{3})
    (\s|,)
    )''', re.VERBOSE)
phone_v6 = re.compile(r'''(
    \D
    \s
    (\d{4}|\(\d{4}\))
    (\s|-)
    (\d{2})
    (\s|-)
    (\d{2})
    (\s|-)
    (\d{2})
    (\s|,)
    )''', re.VERBOSE)
phone_v7 = re.compile(r'''(
    \D
    (\d{12})
    )''', re.VERBOSE)
phone_v8 = re.compile(r'''(
    (\d)
    (\s|-|\.)?          # delimiter
    (\d{3}|\(\d{3}\))  # region code
    (\s|-|\.)           # delimiter
    (\d)
    (\s|-|\.)           # delimiter
    (\d{2})
    (\d)
    (\s|-|\.)           # delimiter
    (\d)
    (\d{2})
    )''', re.VERBOSE)
#  '219-449' - is a number, some 6-digit numbers are salary "130-000"
# 3-752-983-16-83 -  incorrect (correct 375298316831), but if you add \D or \s the number at the end of the line will
# not be found
wb = openpyxl.load_workbook('vacancies.xlsx')

sheet = wb['Лист1']
phone_numbers = []

for row in range(1, sheet.max_row):
    for groups in phone_v1.findall(sheet.cell(row=row, column=1).value):
        phone_num = '-'.join([groups[1], groups[3], groups[5], groups[7], groups[9]])
        if phone_num not in phone_numbers:
            phone_numbers.append(phone_num)
    for groups in phone_v2.findall(sheet.cell(row=row, column=1).value):
        phone_num = '-'.join([groups[1], groups[3], groups[5]])
        if phone_num not in phone_numbers:
            phone_numbers.append(phone_num)
    for groups in phone_v3.findall(sheet.cell(row=row, column=1).value):
        phone_num = '-'.join([groups[2], groups[4]])
        if phone_num not in phone_numbers:
            phone_numbers.append(phone_num)
    for groups in phone_v4.findall(sheet.cell(row=row, column=1).value):
        phone_num = '-'.join([groups[1], groups[3], groups[5], groups[7]])
        if phone_num not in phone_numbers:
            phone_numbers.append(phone_num)
    for groups in phone_v5.findall(sheet.cell(row=row, column=1).value):
        phone_num = '-'.join([groups[1], groups[3], groups[5], groups[7]])
        if phone_num not in phone_numbers:
            phone_numbers.append(phone_num)
    for groups in phone_v6.findall(sheet.cell(row=row, column=1).value):
        phone_num = '-'.join([groups[1], groups[3], groups[5], groups[7]])
        if phone_num not in phone_numbers:
            phone_numbers.append(phone_num)
    for groups in phone_v7.findall(sheet.cell(row=row, column=1).value):
        phone_num = ''.join([groups[1]])
        if phone_num not in phone_numbers:
            phone_numbers.append(phone_num)
    for groups in phone_v8.findall(sheet.cell(row=row, column=1).value):
        phone_num = ''.join([groups[0]])
        if phone_num not in phone_numbers:
            phone_numbers.append(phone_num)

wb.create_sheet(title='Phone numbers in order')
table_of_numbers = wb['Phone numbers in order']
for i, line in enumerate(phone_numbers, start=1):
    table_of_numbers['A' + str(i)] = line

wb.save('update.xlsx')
